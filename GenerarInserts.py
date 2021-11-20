# -*- coding: utf-8 -*-
"""
Created on Tue Nov 16 13:50:45 2021

@author: fabri
"""

import openpyxl

file = openpyxl.load_workbook('NormalizacionReporte.xlsx',read_only=True)
# productos = file.get_sheet_by_name('Data_Productos')
# insert1 = []

# for i in range(2,productos.max_row+1):
#     insert1.append(f"""insert into producto values(
# {productos.cell(i, 6).value},
# {productos.cell(i, 1).value},
# '{productos.cell(i, 3).value}',
# {productos.cell(i, 5).value},
# '{productos.cell(i, 2).value}',
# '{productos.cell(i, 4).value}');
# """)

# clientes = file.get_sheet_by_name('Data_Cliente')
# insert1 = []

# for i in range(2,clientes.max_row+1):
#     insert1.append(f"""insert into usuario values(
# {clientes.cell(i, 1).value},
# '{clientes.cell(i, 2).value}',
# '{clientes.cell(i, 3).value}',
# 'Prueba123');
# """)


# vendedores = file.get_sheet_by_name('Data_Vendedor')
# insert1 = []

# for i in range(2,vendedores.max_row+1):
#     insert1.append(f"""insert into vendedor values(
# {vendedores.cell(i, 1).value},
# '{vendedores.cell(i, 2).value}');
# """)

pedidos = file.get_sheet_by_name('Data_Orden')
insert1 = []
archivo = open("Pedidos.txt","w")

for i in range(2,pedidos.max_row+1):
    fecha = ""
    dia = str(pedidos.cell(i, 8).value)
    mes = str(pedidos.cell(i, 9).value)
    año = str(pedidos.cell(i, 10).value)
    if len(dia) == 1:
        fecha += '0'
        fecha += dia
        fecha += '/'
    else:
        fecha += dia
        fecha += '/'
    if len(mes) == 1:
        fecha += '0'
        fecha += mes
        fecha += '/'
        fecha += año
    else:
        fecha += mes
        fecha += '/'
        fecha += año
    insert1.append(f"""insert into pedido values(
{pedidos.cell(i, 1).value},
'{pedidos.cell(i, 4).value}',
'{fecha}',
'{pedidos.cell(i, 3).value}',
{pedidos.cell(i, 7).value},
{pedidos.cell(i, 6).value},
{pedidos.cell(i, 5).value});
""")
for c in insert1:
    #print(c)
    archivo.write(f"{c}\n") 

# pedidos = file.get_sheet_by_name('Data_Categoria')
# insert1 = []

# for i in range(2,pedidos.max_row+1):
#     insert1.append(f"""insert into categoria values(
# '{pedidos.cell(i, 2).value.upper()}',
# '{pedidos.cell(i, 1).value}');
# """)

# for c in insert1:
#     print(c) 
    
#DATA ALMACEN
# pedidos = file.get_sheet_by_name('Data_Almacen')
# insert1 = []

# for i in range(2,pedidos.max_row+1):
#     insert1.append(f"""insert into categoria values(
# '{pedidos.cell(i, 1).value}',
# '{pedidos.cell(i, 2).value}');
# """)

#Data STOCK
# pedidos = file.get_sheet_by_name('DataStock')
# insert1 = []

# for i in range(2,pedidos.max_row+1):
#     insert1.append(f"""insert into stock values(
# '{pedidos.cell(i, 1).value}',
# {pedidos.cell(i, 2).value},
# {pedidos.cell(i, 3).value});
# """)

# for c in insert1:
#     print(c) 
  
# pedidos = file.get_sheet_by_name('Data_Detalle')
# insert1 = []
# archivo = open("Detalles.txt","w")

# for i in range(2,pedidos.max_row+1):
#     insert1.append(f"""insert into detalle values(
# {pedidos.cell(i, 4).value},
# {pedidos.cell(i, 3).value},
# {pedidos.cell(i, 1).value},
# {pedidos.cell(i, 2).value},
# {pedidos.cell(i, 5).value});
# """)
# for c in insert1:
#     archivo.write(f"{c}\n") 

archivo.close()
file.close()







